from openpyxl import load_workbook
import os
import pandas as pd

# excel 파일들의 리스트
pest_dir = "dataset/pest"

def load_csv():
    df_leaf = pd.read_csv(os.path.join(pest_dir, 'ep_leaf.csv'))
    df_stem = pd.read_csv(os.path.join(pest_dir, 'ep_stem.csv'))
    df_note = pd.read_csv(os.path.join(pest_dir, 'ep_note.csv'))

    df_leaf = df_leaf.replace(to_replace='A', value=88)
    df_leaf = df_leaf.replace(to_replace='B', value=63)
    df_leaf = df_leaf.replace(to_replace='C', value=37)
    df_leaf = df_leaf.replace(to_replace='D', value=13)

    df_leaf = df_leaf.drop(['sample_l', 'ln'], axis=1)
    df_leaf = df_leaf.groupby(['farm', 'crop', 'date']).agg(['count', 'sum'])
    df_leaf = df_leaf.swaplevel(0, 1, axis=1)
    df_leaf = df_leaf.reindex(sorted(df_leaf.columns), axis=1)
    df_leaf.to_csv(os.path.join(pest_dir, "ep_leaf_proc.csv"), encoding='utf-8-sig')

    df_stem = df_stem.drop(['sample_s'], axis=1)
    df_stem = df_stem.groupby(['farm', 'crop', 'date']).agg(['count', 'sum'])
    df_stem = df_stem.swaplevel(0, 1, axis=1)
    df_stem = df_stem.reindex(sorted(df_stem.columns), axis=1)
    df_stem.to_csv(os.path.join(pest_dir, "ep_stem_proc.csv"), encoding='utf-8-sig')

    df_note = df_note.groupby(['farm', 'crop', 'date']).first()


    df = pd.concat([df_leaf, df_stem, df_note], axis=1)
    df.to_csv(os.path.join(pest_dir, "ep_proc.csv"), encoding='utf-8-sig')

    df_tomato = df[df.index.get_level_values(1) == '방울토마토']
    df_tomato.to_csv(os.path.join(pest_dir, "ep_tomato.csv"), encoding='utf-8-sig')
    df_berry = df[df.index.get_level_values(1) == '딸기']
    df_berry.to_csv(os.path.join(pest_dir, "ep_berry.csv"), encoding='utf-8-sig')

    df = pd.concat([df_leaf, df_stem], axis=1)
    return df


def parse():
    files = os.listdir(pest_dir)
    leaf_data = []
    stem_data = []
    note_data = []

    def refine_value(val):
        if isinstance(val, str) and not val.isnumeric():
            val = val[0]
            if (val.upper() == 'E' or val.upper() == 'X' or val == '-'):
                val = None
            if val != None and val.upper() == 'O':
                val = 1
        if val == 0 or val == '0':
            val = None
        return val
    
    def refine_note(val:str):
        if val == None:
            return None
        
        if val.strip() == '-':
            return None

        if val.startswith("-"):
            val = " " + val

        return val

    for file in files:
        file2 = file.replace('.', '-')
        splited = file2.split("-")
        farm = splited[1]
        crop = splited[2]
        date = splited[3]

        print(farm, crop, date)
        if file.startswith('~'): # 열려서 임시로 만들어진 파일
            continue

        if crop == '방울토마토':
            l_d = ['잎곰팡이', '잿빛곰팡이', '흰가루병', '균핵병']
            s_d = ['줄기역병', '풋마름병', '세균성궤양병', '시들음병', 'TYLCV', 'TSWV', 'ToB', 'ToCV']
            l_i = ['진딧물', '응애', '차먼지응애', '녹응애', '담배가루이', '온실가루이', '총채벌레', '잎굴파리']
            s_i = ['나방류']
        else: # 딸기
            l_d = ['흰가루병', '잿빛곰팡이', '세균성모무늬병', '뱀눈무늬병']
            s_d = ['탄저병', '위황병', '역병', '눈마름병', '점균병']
            l_i = ['진딧물', '응애', '차먼지응애', '가루이류', '총채벌레']
            s_i = ['파밤나방류', '거세미나방', '딸기잎벌레']

        file_path = os.path.join(pest_dir, file)
        wb = load_workbook(file_path)
        ws = wb.active

        # 잎
        for si in range(20):
            for li in range(5):
                leaf = {}
                leaf['farm'] = farm
                leaf['crop'] = crop
                leaf['date'] = date
                leaf['sample_l'] = 'sample_l_' + str(si + 1)
                leaf['ln'] = 'L' + str(li + 1)

                # 병든 잎률
                base_row = 6
                base_col = 3
                for di, desease in enumerate(l_d):
                    row = (base_row if si < 10 else base_row + len(l_d) + 2) + di
                    sii = si if si < 10 else si - 10
                    col = base_col + sii * 5 + li
                    val = ws.cell(row, col).value
                    val = refine_value(val)
                    leaf[desease] = val
                
                # 해충 잎
                base_row = 35 if crop == "방울토마토" else 30
                for ii, pest in enumerate(l_i):
                    if crop == '방울토마토':
                        row = (base_row if si < 10 else base_row + (len(l_i) - 1) + 2) + ii
                    else:
                        row = (base_row if si < 10 else base_row + len(l_i) + 2) + ii
                    sii = si if si < 10 else si - 10
                    col = base_col + sii * 5 + li
                    if pest == '잎굴파리':
                        row = 53 if si < 10 else 56

                    val = ws.cell(row, col).value
                    if (val == 'A'): val = 4
                    if (val == 'B'): val = 3
                    if (val == 'C'): val = 2
                    if (val == 'D'): val = 1
                    val = refine_value(val)
                    leaf[pest] = val

                leaf_data.append(leaf)


        # 포기(줄기)
        for si in range(50):
            stem = {}
            stem['farm'] = farm
            stem['crop'] = crop
            stem['date'] = date
            stem['sample_s'] = 'sample_s_' + str(si + 1)

            # 병든 포기율
            base_row = 19 
            base_col = 3
            for di, desease in enumerate(s_d):
                row = base_row + di
                if crop == '방울토마토':
                    row = (base_row if di < 4 else base_row + 2) + di
                col = base_col + si 
                val = ws.cell(row, col).value
                val = refine_value(val)
                if (val == 'D'): val = 1
                stem[desease] = val
            
            
            # 포기당 마리수 (나방)
            base_row = 59 if crop == "방울토마토" else 44
            if (si < 20):
                for ii, pest in enumerate(s_i):
                    row = base_row + ii
                    col = base_col + si * 2 
                    val = ws.cell(row, col).value
                    val = refine_value(val)
                    stem[pest] = val

            
            stem_data.append(stem)

        if crop == '방울토마토':
            control = ws['A64'].value
            job = ws['N64'].value
            note = ws['AJ64'].value
        else:
            control = ws['A51'].value
            job = ws['N51'].value
            note = ws['AJ51'].value

        control = refine_note(control)
        job = refine_note(job)
        note = refine_note(note)

        note_data.append({'farm': farm, 'crop': crop, 'date': date, 'control': control, 'job': job, 'note': note})


    df_leaf = pd.DataFrame(leaf_data)
    df_leaf = df_leaf.set_index(['farm', 'crop', 'date', 'sample_l', 'ln'])
    df_leaf.to_csv(os.path.join(pest_dir, "ep_leaf.csv"), encoding='utf-8-sig')

    df_stem = pd.DataFrame(stem_data)
    df_stem = df_stem.set_index(['farm', 'crop', 'date', 'sample_s'])
    df_stem.to_csv(os.path.join(pest_dir, "ep_stem.csv"), encoding='utf-8-sig')

    df_note = pd.DataFrame(note_data)
    df_note = df_note.set_index(['farm', 'crop', 'date'])
    df_note.to_csv(os.path.join(pest_dir, "ep_note.csv"), encoding='utf-8-sig')


    df_leaf_drop = df_leaf.dropna(how='all').dropna(how='all', axis=1)
    df_leaf_drop.to_csv(os.path.join(pest_dir, "ep_leaf_drop.csv"), encoding='utf-8-sig')
    df_stem_drop = df_stem.dropna(how='all').dropna(how='all', axis=1)
    df_stem_drop.to_csv(os.path.join(pest_dir, "ep_stem_drop.csv"), encoding='utf-8-sig')
    df_note_drop = df_note.dropna(how='all')
    df_note_drop.to_csv(os.path.join(pest_dir, "ep_note_drop.csv"), encoding='utf-8-sig')


if __name__ == '__main__':
    load_csv()
    #parse()
